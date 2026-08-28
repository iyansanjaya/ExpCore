import os
import sys
import glob
import re
import csv
import unicodedata
import pdfplumber
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ==========================================
# KONFIGURASI TEMA
# ==========================================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class ExpCore(ctk.CTk):
    """Aplikasi desktop untuk mengekstrak data PDF Coretax ke Excel."""

    # ── Palet Warna ──
    # Terinspirasi oleh estetika Linear / Raycast — minimal, muted, satu aksen.
    C = {
        "bg":            "#0f1117",
        "surface":       "#161820",
        "surface_hover": "#1c1e28",
        "border":        "#242630",
        "border_focus":  "#3b3e50",

        "accent":        "#818cf8",
        "accent_hover":  "#727de6",
        "accent_muted":  "#272a48",

        "teal":          "#5eead4",
        "teal_hover":    "#4cd9c3",
        "teal_muted":    "#1a3a38",

        "green":         "#4ade80",
        "red":           "#f87171",
        "amber":         "#fbbf24",

        "text":          "#e2e4eb",
        "text_sub":      "#9295a5",
        "text_muted":    "#50536a",

        "input_bg":      "#11131a",
        "log_bg":        "#0c0e14",
    }

    def __init__(self):
        super().__init__()

        self.title("ExpCore")
        self.geometry("940x620")
        self.minsize(840, 540)
        self.configure(fg_color=self.C["bg"])

        # ── Icon ──
        # Set AppUserModelID agar Windows menampilkan ikon ExpCore di taskbar,
        # bukan ikon Python/Nuitka default.
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("iyansanjaya.expcore.1.0")
        except Exception:
            pass

        if "__compiled__" in dir():
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))

        # iconbitmap (.ico) — untuk title bar & taskbar
        ico_path = os.path.join(base_dir, "icon.ico")
        if os.path.exists(ico_path):
            self.after(200, lambda: self.iconbitmap(ico_path))

        # iconphoto (.png) — resolusi lebih tinggi untuk alt-tab & taskbar
        png_path = os.path.join(base_dir, "icon.png")
        if os.path.exists(png_path):
            from tkinter import PhotoImage
            self._icon_img = PhotoImage(file=png_path)
            self.iconphoto(True, self._icon_img)

        # ── State ──
        self.folder_path_bupot = ctk.StringVar(value="")
        self.folder_path_pm = ctk.StringVar(value="")
        self.folder_path_bupot2024 = ctk.StringVar(value="")
        self.folder_path_rename = ctk.StringVar(value="")
        self._anim_id = None

        self._build()

    # ══════════════════════════════════════════
    #  LAYOUT
    # ══════════════════════════════════════════
    def _build(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_content()
        self._navigate("bupot")

    # ──────────────────────────────────────────
    #  Sidebar
    # ──────────────────────────────────────────
    def _build_sidebar(self):
        self.sidebar = ctk.CTkFrame(
            self, width=210, corner_radius=0,
            fg_color=self.C["bg"],
            border_width=0,
        )
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_columnconfigure(0, weight=1)
        self.sidebar.grid_rowconfigure(7, weight=1)

        # ── Brand ──
        ctk.CTkLabel(
            self.sidebar, text="ExpCore",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=self.C["text"],
        ).grid(row=0, column=0, padx=24, pady=(30, 2), sticky="w")

        ctk.CTkLabel(
            self.sidebar, text="by Iyan Sanjaya",
            font=ctk.CTkFont(size=11),
            text_color=self.C["text_muted"],
        ).grid(row=1, column=0, padx=24, pady=(0, 24), sticky="w")

        # ── Navigation ──
        self.nav = {}
        self.nav["bupot"] = self._make_nav(
            self.sidebar, "Bukti Potong 2026", 2,
            lambda: self._navigate("bupot"),
        )
        self.nav["bupot2024"] = self._make_nav(
            self.sidebar, "Bukti Potong 2024", 3,
            lambda: self._navigate("bupot2024"),
        )
        self.nav["pm"] = self._make_nav(
            self.sidebar, "Pajak Masukan", 4,
            lambda: self._navigate("pm"),
        )
        self.nav["rename"] = self._make_nav(
            self.sidebar, "Penamaan Bupot", 5,
            lambda: self._navigate("rename"),
        )

        # ── Divider ──
        ctk.CTkFrame(
            self.sidebar, height=1, fg_color=self.C["border"],
        ).grid(row=6, column=0, padx=20, pady=(20, 0), sticky="ew")

        # ── Version ──
        ctk.CTkLabel(
            self.sidebar, text="v1.4",
            font=ctk.CTkFont(size=10),
            text_color=self.C["text_muted"],
        ).grid(row=8, column=0, padx=24, pady=(0, 20), sticky="sw")

    def _make_nav(self, parent, label, row, cmd):
        """Membuat tombol navigasi sidebar — minimalis, tanpa ikon."""
        btn = ctk.CTkButton(
            parent, text=label, anchor="w",
            height=38, corner_radius=8,
            font=ctk.CTkFont(size=13),
            fg_color="transparent",
            text_color=self.C["text_sub"],
            hover_color=self.C["surface_hover"],
            command=cmd,
        )
        btn.grid(row=row, column=0, padx=14, pady=1, sticky="ew")
        return btn

    def _navigate(self, key):
        """Switch halaman & update state sidebar."""
        for name, btn in self.nav.items():
            if name == key:
                btn.configure(
                    fg_color=self.C["accent_muted"],
                    text_color=self.C["accent"],
                    hover_color=self.C["accent_muted"],
                )
            else:
                btn.configure(
                    fg_color="transparent",
                    text_color=self.C["text_sub"],
                    hover_color=self.C["surface_hover"],
                )
        for name, page in self.pages.items():
            if name == key:
                page.grid(row=0, column=0, sticky="nsew", padx=(20, 28), pady=28)
            else:
                page.grid_forget()

    # ──────────────────────────────────────────
    #  Content wrapper
    # ──────────────────────────────────────────
    def _build_content(self):
        # Garis vertikal tipis sebagai pembatas sidebar — content
        divider = ctk.CTkFrame(self, width=1, fg_color=self.C["border"], corner_radius=0)
        divider.grid(row=0, column=0, sticky="nse")

        self.content = ctk.CTkFrame(self, fg_color=self.C["bg"], corner_radius=0)
        self.content.grid(row=0, column=1, sticky="nsew")
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(0, weight=1)

        self.pages = {}
        self._page_bupot()
        self._page_pm()
        self._page_bupot_2024()
        self._page_rename_bupot()

    # ══════════════════════════════════════════
    #  HALAMAN — BUKTI POTONG
    # ══════════════════════════════════════════
    def _page_bupot(self):
        p = self._page_frame()
        self.pages["bupot"] = p

        # Header
        self._heading(p, "Bukti Potong 2026", "Ekstrak data PDF Bukti Potong Coretax ke Excel.", row=0)

        # Folder picker
        pick = self._picker_frame(p, row=1)
        self.entry_bupot = self._folder_entry(pick, self.folder_path_bupot)
        self._browse_btn(pick, lambda: self.browse_folder(self.folder_path_bupot, self.log_bupot))

        # Log
        log_wrap = self._log_frame(p, row=2)
        self.log_bupot_box = self._log_box(log_wrap)

        # Action
        self.btn_process_bupot = ctk.CTkButton(
            p, text="Mulai Ekstrak", height=44, corner_radius=8,
            fg_color=self.C["accent"], hover_color=self.C["accent_hover"],
            text_color="#ffffff",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.process_bupot,
        )
        self.btn_process_bupot.grid(row=3, column=0, sticky="ew", pady=(12, 0))

    # ══════════════════════════════════════════
    #  HALAMAN — PAJAK MASUKAN
    # ══════════════════════════════════════════
    def _page_pm(self):
        p = self._page_frame()
        self.pages["pm"] = p

        self._heading(p, "Pajak Masukan", "Ekstrak data PDF Faktur Pajak Masukan Coretax ke Excel.", row=0)

        pick = self._picker_frame(p, row=1)
        self.entry_pm = self._folder_entry(pick, self.folder_path_pm)
        self._browse_btn(pick, lambda: self.browse_folder(self.folder_path_pm, self.log_pm))

        log_wrap = self._log_frame(p, row=2)
        self.log_pm_box = self._log_box(log_wrap)

        self.btn_process_pm = ctk.CTkButton(
            p, text="Mulai Ekstrak", height=44, corner_radius=8,
            fg_color=self.C["teal"], hover_color=self.C["teal_hover"],
            text_color="#0f1117",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.process_pm,
        )
        self.btn_process_pm.grid(row=3, column=0, sticky="ew", pady=(12, 0))

    # ══════════════════════════════════════════
    #  HALAMAN — BUKTI POTONG 2024 (formulir BPBS)
    # ══════════════════════════════════════════
    def _page_bupot_2024(self):
        p = self._page_frame()
        self.pages["bupot2024"] = p

        self._heading(
            p, "Bukti Potong 2024",
            "Ekstrak data PDF Bukti Potong formulir BPBS (pra-Coretax) ke Excel.",
            row=0,
        )

        pick = self._picker_frame(p, row=1)
        self.entry_bupot2024 = self._folder_entry(pick, self.folder_path_bupot2024)
        self._browse_btn(pick, lambda: self.browse_folder(self.folder_path_bupot2024, self.log_bupot2024))

        log_wrap = self._log_frame(p, row=2)
        self.log_bupot2024_box = self._log_box(log_wrap)

        self.btn_process_bupot2024 = ctk.CTkButton(
            p, text="Mulai Ekstrak", height=44, corner_radius=8,
            fg_color=self.C["amber"], hover_color="#e0a91f",
            text_color="#0f1117",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.process_bupot_2024,
        )
        self.btn_process_bupot2024.grid(row=3, column=0, sticky="ew", pady=(12, 0))

    def _page_rename_bupot(self):
        p = self._page_frame()
        self.pages["rename"] = p

        self._heading(
            p, "Penamaan Otomatis Bupot",
            "Pratinjau lalu ganti nama PDF berdasarkan identitas Bukti Potong.",
            row=0,
        )

        pick = self._picker_frame(p, row=1)
        self.entry_rename = self._folder_entry(pick, self.folder_path_rename)
        self._browse_btn(pick, lambda: self.browse_folder(self.folder_path_rename, self.log_rename))

        log_wrap = self._log_frame(p, row=2)
        self.log_rename_box = self._log_box(log_wrap)

        actions = ctk.CTkFrame(p, fg_color="transparent")
        actions.grid(row=3, column=0, sticky="ew", pady=(12, 0))
        actions.grid_columnconfigure(0, weight=1)
        actions.grid_columnconfigure(1, weight=1)

        self.btn_preview_rename = ctk.CTkButton(
            actions, text="Pratinjau Nama", height=44, corner_radius=8,
            fg_color=self.C["surface_hover"], hover_color=self.C["border_focus"],
            border_width=1, border_color=self.C["border"],
            text_color=self.C["text"], font=ctk.CTkFont(size=14, weight="bold"),
            command=lambda: self.process_rename_bupot(apply_changes=False),
        )
        self.btn_preview_rename.grid(row=0, column=0, sticky="ew", padx=(0, 6))

        self.btn_apply_rename = ctk.CTkButton(
            actions, text="Terapkan Nama", height=44, corner_radius=8,
            fg_color=self.C["accent"], hover_color=self.C["accent_hover"],
            text_color="#ffffff", font=ctk.CTkFont(size=14, weight="bold"),
            command=lambda: self.process_rename_bupot(apply_changes=True),
        )
        self.btn_apply_rename.grid(row=0, column=1, sticky="ew", padx=(6, 0))

    # ══════════════════════════════════════════
    #  KOMPONEN UI (reusable, bersih)
    # ══════════════════════════════════════════
    def _page_frame(self):
        f = ctk.CTkFrame(self.content, fg_color="transparent")
        f.grid_columnconfigure(0, weight=1)
        f.grid_rowconfigure(2, weight=1)
        return f

    def _heading(self, parent, title, subtitle, row):
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.grid(row=row, column=0, sticky="ew", pady=(0, 20))
        ctk.CTkLabel(
            wrap, text=title,
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=self.C["text"],
        ).pack(anchor="w")
        ctk.CTkLabel(
            wrap, text=subtitle,
            font=ctk.CTkFont(size=13),
            text_color=self.C["text_sub"],
        ).pack(anchor="w", pady=(4, 0))

    def _picker_frame(self, parent, row):
        f = ctk.CTkFrame(
            parent, fg_color=self.C["surface"],
            corner_radius=10, border_width=1,
            border_color=self.C["border"],
        )
        f.grid(row=row, column=0, sticky="ew", pady=(0, 10))
        f.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            f, text="Folder Induk",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=self.C["text_sub"],
        ).grid(row=0, column=0, columnspan=2, padx=16, pady=(14, 6), sticky="w")

        return f

    def _folder_entry(self, parent, var):
        e = ctk.CTkEntry(
            parent, textvariable=var, state="readonly",
            placeholder_text="Belum ada folder induk dipilih",
            height=40, corner_radius=8,
            fg_color=self.C["input_bg"],
            border_color=self.C["border"],
            border_width=1,
            text_color=self.C["text"],
            placeholder_text_color=self.C["text_muted"],
            font=ctk.CTkFont(size=13),
        )
        e.grid(row=1, column=0, padx=(16, 8), pady=(0, 14), sticky="ew")
        return e

    def _browse_btn(self, parent, cmd):
        ctk.CTkButton(
            parent, text="Pilih", width=80, height=40,
            corner_radius=8,
            fg_color=self.C["surface_hover"],
            hover_color=self.C["border_focus"],
            border_width=1, border_color=self.C["border"],
            text_color=self.C["text_sub"],
            font=ctk.CTkFont(size=13),
            command=cmd,
        ).grid(row=1, column=1, padx=(0, 16), pady=(0, 14))

    def _log_frame(self, parent, row):
        f = ctk.CTkFrame(
            parent, fg_color=self.C["surface"],
            corner_radius=10, border_width=1,
            border_color=self.C["border"],
        )
        f.grid(row=row, column=0, sticky="nsew")
        f.grid_columnconfigure(0, weight=1)
        f.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(
            f, text="Log",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=self.C["text_sub"],
        ).grid(row=0, column=0, padx=16, pady=(14, 6), sticky="w")

        return f

    def _log_box(self, parent):
        tb = ctk.CTkTextbox(
            parent, corner_radius=6,
            fg_color=self.C["log_bg"],
            text_color=self.C["text_muted"],
            border_width=1, border_color=self.C["border"],
            font=ctk.CTkFont(family="Consolas", size=12),
            scrollbar_button_color=self.C["border"],
            scrollbar_button_hover_color=self.C["border_focus"],
        )
        tb.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="nsew")
        tb.insert("0.0", f"{self._ts()}  Siap.\n")
        tb.configure(state="disabled")
        return tb

    # ══════════════════════════════════════════
    #  UTILITAS
    # ══════════════════════════════════════════
    def _ts(self):
        return datetime.now().strftime("%H:%M:%S")

    def browse_folder(self, var, log_fn):
        path = filedialog.askdirectory()
        if path:
            var.set(path)
            log_fn(f"Folder: {path}")

    def log_bupot(self, msg):
        self._log(self.log_bupot_box, msg)

    def log_pm(self, msg):
        self._log(self.log_pm_box, msg)

    def log_bupot2024(self, msg):
        self._log(self.log_bupot2024_box, msg)

    def log_rename(self, msg):
        self._log(self.log_rename_box, msg)

    def _log(self, tb, msg):
        tb.configure(state="normal")
        tb.insert("end", f"{self._ts()}  {msg}\n")
        tb.see("end")
        tb.configure(state="disabled")
        self.update_idletasks()

    def _pulse_start(self, btn):
        """Animasi loading minimalis pada tombol."""
        self._anim_dots = 0

        def tick():
            self._anim_dots = (self._anim_dots % 3) + 1
            btn.configure(text="Memproses" + " ." * self._anim_dots)
            self._anim_id = self.after(420, tick)
        tick()

    def _pulse_stop(self, btn, label, color):
        if self._anim_id:
            self.after_cancel(self._anim_id)
            self._anim_id = None
        btn.configure(text=label, state="normal", fg_color=color)

    @staticmethod
    def _write_excel(df, output_path, sheet_name, text_cols=(), money_cols=()):
        """Tulis DataFrame ke Excel dengan header, format angka, dan lebar kolom.

        text_cols dipaksa format teks agar NPWP, masa pajak, dan nomor dokumen
        tidak diubah Excel jadi angka atau tanggal.
        """
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            for cell in ws[1]:
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.font = Font(bold=True)

            col_idx = {col: i + 1 for i, col in enumerate(df.columns)}
            for col_name in text_cols:
                if col_name in col_idx:
                    for cell in ws[get_column_letter(col_idx[col_name])]:
                        cell.number_format = '@'

            for col_name in money_cols:
                if col_name in col_idx:
                    for cell in ws[get_column_letter(col_idx[col_name])][1:]:
                        cell.number_format = '#,##0'

            if "Tarif (%)" in col_idx:
                for cell in ws[get_column_letter(col_idx["Tarif (%)"])][1:]:
                    cell.number_format = '0.00'

            if "No" in col_idx:
                for cell in ws[get_column_letter(col_idx["No"])][1:]:
                    cell.alignment = Alignment(horizontal='center')

            for column in ws.columns:
                max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 55)

    @staticmethod
    def parse_values(val_str):
        v = str(val_str).replace('%', '').replace('Rp', '').strip()
        if ',' in v:
            v = v.replace('.', '').replace(',', '.')
        else:
            v = v.replace('.', '')
        try: return float(v)
        except: return 0.0

    @staticmethod
    def parse_tarif(val_str):
        v = str(val_str).replace('%', '').strip()
        v = v.replace(',', '.')
        try: return float(v)
        except: return 0.0

    @staticmethod
    def _clean_filename_value(value):
        value = unicodedata.normalize("NFKC", str(value or "")).replace("\u00a0", " ")
        return re.sub(r"\s+", " ", value).strip(" :-–—\t\r\n")

    @staticmethod
    def _safe_filename(value, max_len=180):
        value = ExpCore._clean_filename_value(value)
        value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", value)
        value = re.sub(r"\s+", " ", value).rstrip(". ").strip()
        return (value or "UNKNOWN")[:max_len].rstrip(". ")

    @staticmethod
    def _extract_rename_bupot_data(text):
        # ponytail: tambah fallback hanya saat ada contoh layout Coretax yang benar-benar gagal.
        text = unicodedata.normalize("NFKC", text or "").replace("\u00a0", " ")
        text_flat = re.sub(r"\s+", " ", text)
        top_text = text[:2500]
        data = {
            "NAMA_PEMOTONG": "",
            "NOMOR_BUKTI": "",
            "MASA_PAJAK": "",
            "SIFAT": "",
            "STATUS": "",
        }

        name_patterns = [
            r"C\.?\s*3\s*(?:NAMA\s*)?PEMOTONG.*?[:：]\s*(.*?)\s*C\.?\s*4\b",
            r"C\.?\s*3\s*NAMA\s*PEMOTONG[^:：]*[:：]\s*([^\n\r]{2,100})",
            r"NAMA\s*PEMOTONG\s*DAN/?ATAU\s*PEMUNGUT\s*PPh\s*[:：]?\s*([^\n\r]{2,100})",
        ]
        for pattern in name_patterns:
            match = re.search(pattern, text_flat if "C\\.?\\s*4" in pattern else text, re.IGNORECASE)
            if match:
                # Label "…DAN/ATAU PEMUNGUT PPh" terpotong baris, sisa "PPh" ikut tercapture.
                nama = re.sub(r"\s*PPh\s*$", "", match.group(1), flags=re.IGNORECASE)
                data["NAMA_PEMOTONG"] = ExpCore._clean_filename_value(nama)
                break

        header = re.search(
            r"\b([A-Z0-9-]{8,20})\s+(\d{2}-\d{4})\s+"
            r"(TIDAK\s*FINAL|FINAL)\s+"
            r"(NORMAL|PEMBETULAN(?:\s*(?:KE-?)?\s*\d+)?|DIBATALKAN)\b",
            top_text,
            re.IGNORECASE,
        )
        if header:
            data["NOMOR_BUKTI"] = ExpCore._clean_filename_value(header.group(1)).upper()
            data["MASA_PAJAK"] = ExpCore._clean_filename_value(header.group(2)).upper()
            data["SIFAT"] = ExpCore._clean_filename_value(header.group(3)).upper()
            data["STATUS"] = ExpCore._clean_filename_value(header.group(4)).upper()
            return data

        match = re.search(r"\b(\d{2}-\d{4})\b", top_text)
        if match:
            data["MASA_PAJAK"] = match.group(1)

        match = re.search(r"\b(TIDAK\s*FINAL|FINAL)\b", top_text, re.IGNORECASE)
        if match:
            data["SIFAT"] = ExpCore._clean_filename_value(match.group(1)).upper()

        match = re.search(
            r"\b(NORMAL|PEMBETULAN(?:\s*(?:KE-?)?\s*\d+)?|DIBATALKAN)\b",
            top_text,
            re.IGNORECASE,
        )
        if match:
            data["STATUS"] = ExpCore._clean_filename_value(match.group(1)).upper()

        for candidate in re.findall(r"\b[A-Z0-9-]{8,20}\b", top_text.upper()):
            if any(ch.isalpha() for ch in candidate) and any(ch.isdigit() for ch in candidate):
                data["NOMOR_BUKTI"] = candidate
                break

        return data

    @staticmethod
    def _unique_filename(path):
        if not os.path.exists(path):
            return path
        stem, suffix = os.path.splitext(path)
        number = 2
        while os.path.exists(f"{stem} ({number}){suffix}"):
            number += 1
        return f"{stem} ({number}){suffix}"

    def process_rename_bupot(self, apply_changes=False):
        folder = self.folder_path_rename.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("Peringatan", "Silakan pilih folder induk terlebih dahulu!")
            return

        pdf_files = sorted(glob.glob(os.path.join(folder, "**", "*.pdf"), recursive=True))
        if not pdf_files:
            messagebox.showerror("Error", "Tidak ada file PDF di folder atau subfolder tersebut!")
            return

        if apply_changes and not messagebox.askyesno(
            "Konfirmasi",
            f"Ganti nama {len(pdf_files)} PDF yang datanya lengkap?\n\n"
            "PDF dengan data tidak lengkap akan dilewati.",
        ):
            return

        button = self.btn_apply_rename if apply_changes else self.btn_preview_rename
        label = "Terapkan Nama" if apply_changes else "Pratinjau Nama"
        color = self.C["accent"] if apply_changes else self.C["surface_hover"]
        self.btn_preview_rename.configure(state="disabled")
        self.btn_apply_rename.configure(state="disabled")
        self._pulse_start(button)

        mode = "Penerapan" if apply_changes else "Pratinjau"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = os.path.join(folder, f"Log_Penamaan_Bupot_{mode}_{timestamp}.csv")
        fields = [
            "status", "folder_sumber", "nama_lama", "nama_baru", "data_tidak_lengkap",
            "NAMA_PEMOTONG", "NOMOR_BUKTI", "MASA_PAJAK", "SIFAT", "STATUS",
        ]
        complete = skipped = failed = unchanged = 0
        self.log_rename(f"{mode}: memeriksa {len(pdf_files)} PDF …")

        try:
            with open(log_path, "w", newline="", encoding="utf-8-sig") as log_file:
                writer = csv.DictWriter(log_file, fieldnames=fields)
                writer.writeheader()

                for file_pdf in pdf_files:
                    relative_file = os.path.relpath(file_pdf, folder)
                    folder_source = os.path.dirname(relative_file) or "."
                    row = {
                        "status": "ERROR",
                        "folder_sumber": folder_source,
                        "nama_lama": os.path.basename(file_pdf),
                        "nama_baru": "",
                        "data_tidak_lengkap": "exception",
                    }
                    try:
                        with pdfplumber.open(file_pdf) as pdf:
                            page_texts = []
                            for page in pdf.pages:
                                page_text = page.extract_text()
                                if page_text:
                                    page_texts.append(page_text)

                        data = self._extract_rename_bupot_data("\n".join(page_texts))
                        missing = [key for key, value in data.items() if not value]
                        components = [
                            self._safe_filename(data["NAMA_PEMOTONG"] or "UNKNOWN_NAMA", 80),
                            self._safe_filename(data["NOMOR_BUKTI"] or "UNKNOWN_NOMOR", 30),
                            self._safe_filename(data["MASA_PAJAK"] or "UNKNOWN_MASA", 20),
                            self._safe_filename(data["SIFAT"] or "UNKNOWN_SIFAT", 20),
                            self._safe_filename(data["STATUS"] or "UNKNOWN_STATUS", 30),
                        ]
                        new_name = f"{self._safe_filename(' - '.join(components))}.pdf"
                        new_path = os.path.join(os.path.dirname(file_pdf), new_name)

                        if missing:
                            status = "DILEWATI" if apply_changes else "PERLU CEK"
                            skipped += 1
                        elif os.path.basename(file_pdf).casefold() == new_name.casefold():
                            status = "SUDAH SESUAI"
                            unchanged += 1
                        elif apply_changes:
                            new_path = self._unique_filename(new_path)
                            os.rename(file_pdf, new_path)
                            new_name = os.path.basename(new_path)
                            status = "BERHASIL"
                            complete += 1
                        else:
                            new_path = self._unique_filename(new_path)
                            new_name = os.path.basename(new_path)
                            status = "SIAP"
                            complete += 1

                        row.update({
                            "status": status,
                            "nama_baru": new_name,
                            "data_tidak_lengkap": ", ".join(missing),
                            **data,
                        })
                        self.log_rename(f"{status}: {relative_file} → {new_name}")
                    except Exception as error:
                        failed += 1
                        self.log_rename(f"ERROR: {relative_file} — {error}")

                    writer.writerow(row)
                    log_file.flush()

            summary = (
                f"{mode} selesai — {complete} siap/berhasil, {unchanged} sudah sesuai, "
                f"{skipped} perlu diperiksa, {failed} gagal."
            )
            self.log_rename(summary)
            self.log_rename(f"Log: {log_path}")
            messagebox.showinfo("Selesai", f"{summary}\n\nLog disimpan di:\n{log_path}")
        except Exception as error:
            self.log_rename(f"Error: {error}")
            messagebox.showerror("Error", str(error))
        finally:
            self._pulse_stop(button, label, color)
            self.btn_preview_rename.configure(state="normal")
            self.btn_apply_rename.configure(state="normal")

    # ══════════════════════════════════════════
    #  EKSTRAKSI — BUKTI POTONG
    # ══════════════════════════════════════════
    @classmethod
    def _extract_bupot_rows(cls, teks_lengkap):
        """Ambil semua baris objek pajak dari teks satu PDF Bukti Potong."""
        teks_rata = teks_lengkap.replace('\n', ' ')

        # Nomor bukti, masa pajak & status ada di header — logika sama dengan
        # fitur penamaan otomatis, jadi dipakai ulang.
        header = cls._extract_rename_bupot_data(teks_lengkap)
        nomor_bukti = header["NOMOR_BUKTI"] or "-"
        masa_pajak = header["MASA_PAJAK"] or "-"
        status_bukti = header["STATUS"] or "NORMAL"
        nama_pemotong = header["NAMA_PEMOTONG"] or "-"

        match_nama = re.search(r'A\.2\s*NAMA\s*:\s*(.*?)\s*A\.3', teks_rata)
        nama_penerima = match_nama.group(1).strip() if match_nama else "-"
        match_npwp_penerima = re.search(r'A\.1\s*NPWP\s*/\s*NIK\s*:\s*(\d+)', teks_rata)
        npwp_penerima = match_npwp_penerima.group(1) if match_npwp_penerima else "-"
        match_fasilitas = re.search(r'B\.1\s*Jenis Fasilitas\s*:\s*(.*?)\s*B\.2', teks_rata, re.IGNORECASE)
        jenis_fasilitas = match_fasilitas.group(1).strip() if match_fasilitas else "-"
        match_jpph = re.search(r'B\.2\s*Jenis PPh\s*:\s*(.*?)\s*(?:KODE OBJEK PAJAK|B\.3)', teks_rata, re.IGNORECASE)
        jenis_pph = match_jpph.group(1).strip() if match_jpph else "-"
        # Tanggal dibatasi pola tanggal agar tidak menelan sisa label B.8 yang terpotong baris.
        match_jenis_dok = re.search(
            r'Jenis Dokumen\s*:\s*(.*?)\s*Tanggal\s*:\s*(\d{1,2}\s+\S+\s+\d{4})',
            teks_rata, re.IGNORECASE,
        )
        jenis_dokumen = match_jenis_dok.group(1).strip() if match_jenis_dok else "-"
        tanggal_dokumen = match_jenis_dok.group(2).strip() if match_jenis_dok else "-"
        match_nodok = re.search(r'B\.9\s*Nomor Dokumen\s*:\s*(.*?)\s*B\.10', teks_rata, re.IGNORECASE)
        nomor_dokumen = match_nodok.group(1).strip() if match_nodok else "-"
        match_npwp_pemotong = re.search(r'C\.1\s*NPWP\s*/\s*NIK\s*:\s*([\d]+)', teks_rata)
        npwp_pemotong = match_npwp_pemotong.group(1) if match_npwp_pemotong else "-"
        match_tanggal = re.search(r'C\.4\s*TANGGAL\s*:\s*([A-Za-z0-9\s]+?)\s*C\.5', teks_rata)
        tanggal_bupot = match_tanggal.group(1).strip() if match_tanggal else "-"

        match_blok = re.search(r'B\.7\s*(.*?)\s*B\.8\s*Dokumen', teks_rata, re.IGNORECASE)
        if not match_blok:
            return []

        baris_data = []
        blok_tabel = match_blok.group(1).strip()
        for m in re.finditer(r'(\d{2}-\d{3}-\d{2})\s+(.*?)(?=(?:\d{2}-\d{3}-\d{2})|$)', blok_tabel):
            kode_objek = m.group(1)
            isi_baris = m.group(2).strip()

            pattern_angka = r'(?<!\S)(\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s+(\d+(?:[\.,]\d+)?%?)\s+(\d{1,3}(?:\.\d{3})*(?:,\d+)?)(?!\S)'
            match_angka = re.search(pattern_angka, isi_baris)

            if match_angka:
                dpp_str, tarif_str, pph_str = match_angka.groups()
                objek_pajak_desc = isi_baris[:match_angka.start()] + " " + isi_baris[match_angka.end():]
                objek_pajak_desc = re.sub(r'\s+', ' ', objek_pajak_desc).strip()
            else:
                tokens = isi_baris.split()
                if len(tokens) >= 3:
                    dpp_str, tarif_str, pph_str = tokens[-3], tokens[-2], tokens[-1]
                    objek_pajak_desc = " ".join(tokens[:-3])
                else:
                    dpp_str, tarif_str, pph_str = "0", "0", "0"
                    objek_pajak_desc = isi_baris

            baris_data.append({
                "Nomor Dokumen": nomor_bukti, "Masa Pajak": masa_pajak,
                "NPWP/NIK": npwp_penerima, "Nama": nama_penerima,
                "Status Bukti": status_bukti, "Jenis Fasilitas": jenis_fasilitas,
                "Jenis PPh": jenis_pph, "Kode Objek Pajak": kode_objek, "Objek Pajak": objek_pajak_desc,
                "DPP (Rp)": cls.parse_values(dpp_str), "Tarif (%)": cls.parse_tarif(tarif_str),
                "Pajak Penghasilan (Rp)": cls.parse_values(pph_str), "Jenis Dokumen": jenis_dokumen,
                "Nomor Dokumen Dasar": nomor_dokumen, "Tanggal Dokumen": tanggal_dokumen,
                "NPWP/NIK Pemotong": npwp_pemotong, "Nama Pemotong": nama_pemotong,
                "Tanggal Bukti Potong": tanggal_bupot,
            })

        return baris_data

    def process_bupot(self):
        folder = self.folder_path_bupot.get()
        if not os.path.exists(folder) or not folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return

        pdf_files = sorted(glob.glob(os.path.join(folder, "**", "*.pdf"), recursive=True))
        if not pdf_files:
            messagebox.showerror("Error", "Tidak ada file PDF di folder atau subfolder tersebut!")
            return

        self.btn_process_bupot.configure(state="disabled")
        self._pulse_start(self.btn_process_bupot)
        self.log_bupot(f"Memproses {len(pdf_files)} file …")

        semua_baris_data = []
        try:
            for file_pdf in pdf_files:
                nama_file = os.path.basename(file_pdf)
                folder_sumber = os.path.relpath(os.path.dirname(file_pdf), folder)
                self.log_bupot(f"Membaca {os.path.relpath(file_pdf, folder)}")

                with pdfplumber.open(file_pdf) as pdf:
                    teks_lengkap = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])

                for baris in self._extract_bupot_rows(teks_lengkap):
                    baris["Folder Sumber"] = folder_sumber
                    baris["File Name"] = nama_file
                    semua_baris_data.append(baris)

            if semua_baris_data:
                df = pd.DataFrame(semua_baris_data)
                df.insert(0, 'No', range(1, len(df) + 1))
                output_path = os.path.join(folder, "!Hasil_Rekap_Bupot.xlsx")

                self._write_excel(
                    df, output_path, "Rekap",
                    text_cols=["Nomor Dokumen", "Masa Pajak", "NPWP/NIK", "Kode Objek Pajak",
                               "Nomor Dokumen Dasar", "NPWP/NIK Pemotong"],
                    money_cols=["DPP (Rp)", "Pajak Penghasilan (Rp)"],
                )

                self.log_bupot(f"Selesai — {output_path}")
                messagebox.showinfo("Selesai", f"Data Bupot berhasil disimpan di:\n{output_path}")
            else:
                self.log_bupot("Tidak ada data yang ditemukan.")

        except Exception as e:
            self.log_bupot(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))
        finally:
            self._pulse_stop(self.btn_process_bupot, "Mulai Ekstrak", self.C["accent"])

    # ══════════════════════════════════════════
    #  EKSTRAKSI — BUKTI POTONG 2024 (formulir BPBS)
    # ══════════════════════════════════════════
    @staticmethod
    def _despace_digits(value):
        """'4 3 6 1 ...' -> '4361...'. Formulir BPBS menulis angka per kotak."""
        return re.sub(r"\s+", "", value or "")

    @staticmethod
    def _form_date(value):
        """'3 1 dd 0 1 mm 2 0 2 4 yyyy' -> '31-01-2024'. Kotak kosong -> ''."""
        match = re.search(r"([\d\s]*?)dd\s*([\d\s]*?)mm\s*([\d\s]*?)yyyy", value or "")
        if not match:
            return ""
        day, month, year = (ExpCore._despace_digits(g) for g in match.groups())
        if not (day and month and year):
            return ""
        return f"{day.zfill(2)}-{month.zfill(2)}-{year}"

    @classmethod
    def _extract_bupot2024_rows(cls, teks_lengkap):
        """Ambil baris objek pajak dari teks satu PDF Bukti Potong formulir BPBS.

        Berbeda dari BPPU Coretax: angka ditulis per kotak (NPWP, tanggal) dan
        sifat Final/Tidak Final ditandai centang "X" di depan labelnya.
        """
        teks = teks_lengkap or ""
        teks_rata = re.sub(r"[ \t]+", " ", teks.replace("\n", " "))

        def ambil(pattern, sumber=None, flags=re.IGNORECASE):
            match = re.search(pattern, teks_rata if sumber is None else sumber, flags)
            return match.group(1).strip() if match else ""

        npwp = cls._despace_digits(ambil(r"A\.1\s*NPWP\s*:\s*([\d\s]*?)\s*A\.2"))
        nik = cls._despace_digits(ambil(r"A\.2\s*NIK\s*:\s*([\d\s]*?)\s*A\.3"))
        nama = ambil(r"A\.3\s*Nama\s*:\s*(.*?)\s*B\.\s*PAJAK")
        objek_pajak = ambil(r"Keterangan Kode Objek Pajak\s*:\s*(.*?)\s*B\.7")
        nomor_dokumen = ambil(r"B\.7.*?Nomor Dokumen\s+(\S+?)\s+Nama Dokumen")
        nama_dokumen = ambil(r"Nama Dokumen\s+(.*?)\s+Tanggal\s")
        tanggal_dokumen = cls._form_date(ambil(r"Nama Dokumen.*?Tanggal\s+(.*?yyyy)"))
        npwp_pemotong = cls._despace_digits(ambil(r"C\.1\s*NPWP\s*:\s*([\d\s]*?)\s*C\.2"))
        nama_pemotong = ambil(r"C\.2\s*Nama Wajib Pajak\s*:\s*(.*?)\s*C\.3")
        tanggal_bukti = cls._form_date(ambil(r"C\.3\s*Tanggal\s*:\s*(.*?yyyy)"))
        # Label C.5 terpotong baris, kata "elektronik" jatuh ke depan C.5.
        penandatangan = ambil(r"C\.4\s*Nama Penandatangan\s*:\s*(.*?)\s*(?:elektronik\s+)?C\.5")

        # Header H semuanya inline pada teks pdfplumber: nomor, pembetulan,
        # dan tanda centang sifat. Centang muncul sebagai "X" antara nomor
        # bagian dan labelnya, mis. "H.5 X PPh Tidak Final".
        nomor_bukti = cls._despace_digits(ambil(r"H\.1\s*NOMOR\s*:\s*([\d\s]*?)\s*H\.\d"))
        pembetulan_str = ambil(r"H\.2\s*Pembetulan\s*Ke-\s*(\d+)")
        pembetulan = int(pembetulan_str) if pembetulan_str.isdigit() else 0
        dibatalkan = bool(ambil(r"H\.3\s*(X)\s*Pembatalan"))
        final = bool(ambil(r"H\.4\s*(X)\s*PPh\s*Final"))
        tidak_final = bool(ambil(r"H\.5\s*(X)\s*PPh\s*Tidak\s*Final"))

        if dibatalkan:
            status_bukti = "DIBATALKAN"
        elif pembetulan:
            status_bukti = f"PEMBETULAN KE-{pembetulan}"
        else:
            status_bukti = "NORMAL"

        # Tepat satu kotak harus tercentang; selain itu jangan menebak.
        sifat = "-" if final == tidak_final else ("FINAL" if final else "TIDAK FINAL")

        baris_data = []
        pola_baris = (
            r"(\d{1,2}-\d{4})\s+(\d{2}-\d{3}-\d{2})\s+([\d.]+,\d{2})\s+"
            r"(?:(X)\s+)?(\d+[.,]\d+)\s+([\d.]+,\d{2})"
        )
        for match in re.finditer(pola_baris, teks):
            masa, kode_objek, dpp, tarif_tinggi, tarif, pph = match.groups()
            bulan, tahun = masa.split("-")
            baris_data.append({
                "Nomor Bukti Potong": nomor_bukti or "-",
                "Pembetulan Ke": pembetulan,
                "Status Bukti": status_bukti,
                "Sifat": sifat,
                "NPWP": npwp or "-",
                "NIK": nik or "-",
                "Nama": nama or "-",
                "Masa Pajak": f"{bulan.zfill(2)}-{tahun}",
                "Kode Objek Pajak": kode_objek,
                "Objek Pajak": objek_pajak or "-",
                "DPP (Rp)": cls.parse_values(dpp),
                "Tarif Lebih Tinggi": "Ya" if tarif_tinggi else "Tidak",
                "Tarif (%)": cls.parse_tarif(tarif),
                "Pajak Penghasilan (Rp)": cls.parse_values(pph),
                "Nomor Dokumen Referensi": nomor_dokumen or "-",
                "Nama Dokumen": nama_dokumen or "-",
                "Tanggal Dokumen": tanggal_dokumen or "-",
                "NPWP Pemotong": npwp_pemotong or "-",
                "Nama Pemotong": nama_pemotong or "-",
                "Tanggal Bukti Potong": tanggal_bukti or "-",
                "Nama Penandatangan": penandatangan or "-",
            })
        return baris_data

    def process_bupot_2024(self):
        folder = self.folder_path_bupot2024.get()
        if not folder or not os.path.exists(folder):
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return

        pdf_files = sorted(glob.glob(os.path.join(folder, "**", "*.pdf"), recursive=True))
        if not pdf_files:
            messagebox.showerror("Error", "Tidak ada file PDF di folder atau subfolder tersebut!")
            return

        self.btn_process_bupot2024.configure(state="disabled")
        self._pulse_start(self.btn_process_bupot2024)
        self.log_bupot2024(f"Memproses {len(pdf_files)} file …")

        semua_baris_data = []
        dilewati = 0
        try:
            for file_pdf in pdf_files:
                nama_file = os.path.basename(file_pdf)
                folder_sumber = os.path.relpath(os.path.dirname(file_pdf), folder)
                relatif = os.path.relpath(file_pdf, folder)
                self.log_bupot2024(f"Membaca {relatif}")

                try:
                    with pdfplumber.open(file_pdf) as pdf:
                        teks_lengkap = "\n".join(
                            p.extract_text() for p in pdf.pages if p.extract_text()
                        )

                    baris_pdf = self._extract_bupot2024_rows(teks_lengkap)
                    if not baris_pdf:
                        dilewati += 1
                        self.log_bupot2024(f"DILEWATI: {relatif} — bukan formulir BPBS?")
                        continue

                    for baris in baris_pdf:
                        baris["Folder Sumber"] = folder_sumber
                        baris["File Name"] = nama_file
                        semua_baris_data.append(baris)
                except Exception as error:
                    dilewati += 1
                    self.log_bupot2024(f"GAGAL: {relatif} — {error}")

            if semua_baris_data:
                df = pd.DataFrame(semua_baris_data)
                df.insert(0, "No", range(1, len(df) + 1))
                output_path = os.path.join(folder, "!Hasil_Rekap_Bupot_2024.xlsx")
                self._write_excel(
                    df, output_path, "Rekap 2024",
                    text_cols=[
                        "Nomor Bukti Potong", "NPWP", "NIK", "Masa Pajak",
                        "Kode Objek Pajak", "Nomor Dokumen Referensi", "NPWP Pemotong",
                    ],
                    money_cols=["DPP (Rp)", "Pajak Penghasilan (Rp)"],
                )
                self.log_bupot2024(
                    f"Selesai — {len(semua_baris_data)} baris, {dilewati} file dilewati."
                )
                self.log_bupot2024(f"Output: {output_path}")
                messagebox.showinfo(
                    "Selesai",
                    f"Data Bupot 2024 berhasil disimpan di:\n{output_path}",
                )
            else:
                self.log_bupot2024("Tidak ada data yang ditemukan.")
                messagebox.showwarning(
                    "Tidak ada data",
                    "Tidak ada PDF formulir BPBS yang bisa dibaca di folder tersebut.",
                )

        except Exception as e:
            self.log_bupot2024(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))
        finally:
            self._pulse_stop(self.btn_process_bupot2024, "Mulai Ekstrak", self.C["amber"])

    # ══════════════════════════════════════════
    #  EKSTRAKSI — PAJAK MASUKAN
    # ══════════════════════════════════════════
    def process_pm(self):
        folder = self.folder_path_pm.get()
        if not os.path.exists(folder) or not folder:
            messagebox.showwarning("Peringatan", "Silakan pilih folder terlebih dahulu!")
            return

        pdf_files = sorted(glob.glob(os.path.join(folder, "**", "*.pdf"), recursive=True))
        if not pdf_files:
            messagebox.showerror("Error", "Tidak ada file PDF di folder atau subfolder tersebut!")
            return

        self.btn_process_pm.configure(state="disabled")
        self._pulse_start(self.btn_process_pm)
        self.log_pm(f"Memproses {len(pdf_files)} file …")

        semua_baris_data = []
        try:
            for file_pdf in pdf_files:
                nama_file = os.path.basename(file_pdf)
                folder_sumber = os.path.relpath(os.path.dirname(file_pdf), folder)
                self.log_pm(f"Membaca {os.path.relpath(file_pdf, folder)}")

                try:
                    with pdfplumber.open(file_pdf) as pdf:
                        teks_lengkap = "".join([page.extract_text() + "\n" for page in pdf.pages if page.extract_text()])

                        nama_pembeli_match = re.search(r'Pembeli Barang Kena Pajak.*?Nama\s*:\s*([^\n]+)', teks_lengkap, re.DOTALL | re.IGNORECASE)
                        nama_pembeli = nama_pembeli_match.group(1).strip() if nama_pembeli_match else "-"

                        pembeli_block = re.search(r'Pembeli Barang Kena Pajak.*?(?P<npwp>\d{15,16})', teks_lengkap.replace('.', '').replace('-', ''), re.DOTALL | re.IGNORECASE)
                        if pembeli_block:
                            npwp_pembeli = pembeli_block.group('npwp')
                        else:
                            npwp_matches = re.findall(r'NPWP\s*:\s*([\d\.\-]+)', teks_lengkap)
                            npwp_pembeli = npwp_matches[-1].replace('.', '').replace('-', '') if npwp_matches else "-"

                        no_seri_match = re.search(r'Kode dan Nomor Seri Faktur Pajak\s*:\s*([\d\-\.]+)', teks_lengkap, re.IGNORECASE)
                        no_seri = no_seri_match.group(1).replace('.', '').replace('-', '') if no_seri_match else "-"

                        no_urut = 1
                        for page in pdf.pages:
                            tabel_halaman = page.extract_tables()
                            for tabel in tabel_halaman:
                                for row in tabel:
                                    if not row or len(row) < 3: continue

                                    kode_col, desc_col = str(row[1] or ""), str(row[2] or "")
                                    if "Nama Barang" in desc_col or not desc_col.strip() or "Harga Jual" in kode_col:
                                        continue

                                    kodes = re.findall(r'\d{6,}', kode_col)
                                    blocks = re.split(r'PPnBM.*?=.*?\n?', desc_col)

                                    for i, block in enumerate(blocks):
                                        if "Rp" not in block: continue

                                        match_harga = re.search(r'Rp\s*([\d\.]+,\d{2})\s*x\s*([\d\.,]+)\s*([A-Za-z]+)', block)
                                        if match_harga:
                                            harga = float(match_harga.group(1).replace('.', '').replace(',', '.'))
                                            qty = float(match_harga.group(2).replace('.', '').replace(',', '.'))
                                            satuan = match_harga.group(3)
                                            potongan_harga = 0

                                            nama_barang = block[:match_harga.start()].replace('\n', ' ').strip()
                                            kode_barang = kodes[i] if i < len(kodes) else (kodes[0] if kodes else "")

                                            dpp = (harga * qty) - potongan_harga
                                            ppn = dpp * 0.12  # Asumsi PPN 12%
                                            netto = dpp + ppn

                                            semua_baris_data.append({
                                                "Nama Pembeli": nama_pembeli, "NPWP Pembeli": npwp_pembeli,
                                                "Kode dan Nomor Seri Faktur Pajak": no_seri, "No": no_urut,
                                                "Kode Barang": kode_barang, "Nama Barang": nama_barang,
                                                "Harga": harga, "Qty": qty, "Satuan": satuan, "Potongan Harga": potongan_harga,
                                                "DPP": dpp, "PPN": ppn, "NETTO": netto,
                                                "Folder Sumber": folder_sumber, "Nama File PDF": nama_file
                                            })
                                            no_urut += 1
                except Exception as inner_e:
                    self.log_pm(f"Gagal: {nama_file} — {str(inner_e)}")

            if semua_baris_data:
                kolom_urutan = ['Nama Pembeli', 'NPWP Pembeli', 'Kode dan Nomor Seri Faktur Pajak', 'No', 'Kode Barang', 'Nama Barang', 'Harga', 'Qty', 'Satuan', 'Potongan Harga', 'DPP', 'PPN', 'NETTO', 'Folder Sumber', 'Nama File PDF']
                df = pd.DataFrame(semua_baris_data, columns=kolom_urutan)
                output_path = os.path.join(folder, "Hasil_Pajak_Masukan.xlsx")

                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Rekap_Faktur")
                    ws = writer.sheets["Rekap_Faktur"]

                    for cell in ws[1]:
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.font = Font(bold=True)

                    for col_letter in ['B', 'C', 'E']:
                        for cell in ws[col_letter]: cell.number_format = '@'

                    col_idx = {col: i+1 for i, col in enumerate(df.columns)}
                    for col_name in ["Harga", "Potongan Harga", "DPP", "PPN", "NETTO"]:
                        if col_name in col_idx:
                            for cell in ws[get_column_letter(col_idx[col_name])][1:]: cell.number_format = '#,##0'

                    for column in ws.columns:
                        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
                        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 55)

                self.log_pm(f"Selesai — {output_path}")
                messagebox.showinfo("Selesai", f"Data Pajak Masukan berhasil disimpan di:\n{output_path}")
            else:
                self.log_pm("Tidak ada data yang ditemukan.")

        except Exception as e:
            self.log_pm(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))
        finally:
            self._pulse_stop(self.btn_process_pm, "Mulai Ekstrak", self.C["teal"])


if __name__ == "__main__":
    app = ExpCore()
    app.mainloop()
